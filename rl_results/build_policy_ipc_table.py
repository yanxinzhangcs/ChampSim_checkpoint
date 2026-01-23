#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class AppSummary:
  app: str
  source_dir: str
  fixed_policies: Tuple[str, ...]
  per_step: Tuple[Mapping[str, Any], ...]


def parse_args() -> argparse.Namespace:
  parser = argparse.ArgumentParser(
      description=(
          "Aggregate rl_results/*/experiment_summary.json into a step x application table.\n"
          "Outputs an .xlsx with multiple sheets plus optional CSV exports."
      )
  )
  parser.add_argument("--results-dir", type=Path, default=Path("rl_results"), help="Directory containing result subfolders")
  parser.add_argument(
      "--out-xlsx",
      type=Path,
      default=Path("rl_results/policy_ipc_table.xlsx"),
      help="Output Excel path",
  )
  parser.add_argument(
      "--out-long-csv",
      type=Path,
      default=Path("rl_results/policy_ipc_long.csv"),
      help="Output long-form CSV path (application, step, action, ipc, ...)",
  )
  parser.add_argument(
      "--skip-csv",
      action="store_true",
      help="Only write the Excel file",
  )
  return parser.parse_args()


def _strip_known_suffixes(name: str) -> str:
  for suffix in (".champsimtrace.xz", ".champsimtrace", ".xz"):
    if name.endswith(suffix):
      return name[: -len(suffix)]
  return name


def infer_application_name(summary: Mapping[str, Any], folder_name: str) -> str:
  trace = summary.get("config", {}).get("trace")
  if isinstance(trace, str) and trace:
    trace_name = _strip_known_suffixes(Path(trace).name)
    return trace_name
  if folder_name.endswith("_combo"):
    return folder_name[: -len("_combo")]
  return folder_name


def application_sort_key(app_name: str) -> Tuple[int, str]:
  prefix = app_name.split(".", 1)[0]
  try:
    return (int(prefix), app_name)
  except ValueError:
    return (10**9, app_name)


def discover_experiment_summaries(results_dir: Path) -> List[Path]:
  if not results_dir.exists():
    raise FileNotFoundError(f"results dir not found: {results_dir}")

  summary_paths: List[Path] = []
  for path in sorted(results_dir.glob("*/experiment_summary.json")):
    folder = path.parent.name
    if folder and folder[0].isdigit():
      summary_paths.append(path)
  if not summary_paths:
    raise FileNotFoundError(f"No experiment_summary.json found under {results_dir} (numeric-prefix folders only).")
  return summary_paths


def load_app_summary(summary_path: Path) -> AppSummary:
  with summary_path.open("r", encoding="utf-8") as handle:
    summary = json.load(handle)

  folder = summary_path.parent.name
  app_name = infer_application_name(summary, folder)

  fixed_policies = tuple(summary.get("fixed_policies") or [])
  per_step_block = summary.get("per_step_comparison") or {}
  per_step_results = per_step_block.get("results") or []
  if not fixed_policies:
    raise ValueError(f"{summary_path}: missing fixed_policies")
  if not isinstance(per_step_results, list) or not per_step_results:
    raise ValueError(f"{summary_path}: missing per_step_comparison.results")

  return AppSummary(
      app=app_name,
      source_dir=folder,
      fixed_policies=fixed_policies,
      per_step=tuple(per_step_results),
  )


def pick_best_action_key(fixed_policies: Iterable[str], per_action_ipc: Mapping[str, Any]) -> Tuple[Optional[str], Optional[float]]:
  best_action: Optional[str] = None
  best_ipc: Optional[float] = None
  for action in fixed_policies:
    ipc_value = per_action_ipc.get(action)
    if ipc_value is None:
      continue
    try:
      ipc = float(ipc_value)
    except (TypeError, ValueError):
      continue
    if best_ipc is None or ipc > best_ipc:
      best_ipc = ipc
      best_action = action
  return best_action, best_ipc


def format_step_cell(fixed_policies: Tuple[str, ...], step_result: Mapping[str, Any]) -> str:
  per_action_ipc = step_result.get("per_action_ipc") or {}
  if not isinstance(per_action_ipc, dict):
    per_action_ipc = {}

  best_action = None
  top_k = step_result.get("top_k")
  if isinstance(top_k, list) and top_k:
    first = top_k[0]
    if isinstance(first, dict):
      best_action = first.get("action")

  best_ipc = step_result.get("best_ipc")
  try:
    best_ipc_f = float(best_ipc) if best_ipc is not None else None
  except (TypeError, ValueError):
    best_ipc_f = None

  if not isinstance(best_action, str) or not best_action:
    best_action, best_ipc_f = pick_best_action_key(fixed_policies, per_action_ipc)

  if best_action is None:
    best_action = "N/A"

  best_ipc_str = "N/A" if best_ipc_f is None else f"{best_ipc_f:.6f}"
  lines = [f"BEST: {best_action} (IPC={best_ipc_str})"]
  for action in fixed_policies:
    ipc_value = per_action_ipc.get(action)
    try:
      ipc = float(ipc_value)
      ipc_str = f"{ipc:.6f}"
    except (TypeError, ValueError):
      ipc_str = "N/A"
    lines.append(f"{action}: {ipc_str}")
  return "\n".join(lines)


def write_matrix_sheet(
    ws,
    apps: List[AppSummary],
    steps: int,
    value_for_cell,
    column_width: float,
) -> None:
  header_font = Font(bold=True)

  ws.cell(row=1, column=1, value="step").font = header_font
  for idx, app in enumerate(apps, start=2):
    ws.cell(row=1, column=idx, value=app.app).font = header_font

  ws.freeze_panes = "B2"
  ws.column_dimensions["A"].width = 6
  for col in range(2, len(apps) + 2):
    ws.column_dimensions[get_column_letter(col)].width = column_width

  wrap = Alignment(wrap_text=True, vertical="top")

  # Build quick lookup: (app_idx -> step_idx -> result)
  app_steps: List[Dict[int, Mapping[str, Any]]] = []
  for app in apps:
    by_step: Dict[int, Mapping[str, Any]] = {}
    for r in app.per_step:
      step = r.get("step")
      if isinstance(step, int):
        by_step[step] = r
    app_steps.append(by_step)

  for step in range(steps):
    ws.cell(row=step + 2, column=1, value=step)
    for app_idx, app in enumerate(apps):
      result = app_steps[app_idx].get(step)
      cell = ws.cell(row=step + 2, column=app_idx + 2, value=value_for_cell(app, result))
      cell.alignment = wrap


def write_long_csv(
    out_csv: Path,
    apps: List[AppSummary],
    steps: int,
    fixed_policies: Tuple[str, ...],
) -> None:
  out_csv.parent.mkdir(parents=True, exist_ok=True)

  # Build quick lookup per app for step results
  app_steps: List[Dict[int, Mapping[str, Any]]] = []
  for app in apps:
    by_step: Dict[int, Mapping[str, Any]] = {}
    for r in app.per_step:
      step = r.get("step")
      if isinstance(step, int):
        by_step[step] = r
    app_steps.append(by_step)

  with out_csv.open("w", encoding="utf-8", newline="") as handle:
    writer = csv.writer(handle)
    writer.writerow(["application", "step", "action", "ipc", "is_best", "best_action", "best_ipc", "base_ipc"])

    for app_idx, app in enumerate(apps):
      for step in range(steps):
        r = app_steps[app_idx].get(step)
        if r is None:
          continue
        per_action_ipc = r.get("per_action_ipc") or {}
        if not isinstance(per_action_ipc, dict):
          per_action_ipc = {}

        best_action, best_ipc = pick_best_action_key(fixed_policies, per_action_ipc)
        base_ipc = r.get("base_ipc")
        for action in fixed_policies:
          ipc_value = per_action_ipc.get(action)
          writer.writerow(
              [
                  app.app,
                  step,
                  action,
                  ipc_value,
                  1 if best_action == action else 0,
                  best_action,
                  best_ipc,
                  base_ipc,
              ]
          )


def main() -> None:
  args = parse_args()

  summary_paths = discover_experiment_summaries(args.results_dir)
  apps = [load_app_summary(p) for p in summary_paths]
  apps.sort(key=lambda a: application_sort_key(a.app))

  fixed_policies = apps[0].fixed_policies
  for app in apps[1:]:
    if app.fixed_policies != fixed_policies:
      raise ValueError(f"fixed_policies mismatch: {apps[0].source_dir} vs {app.source_dir}")

  steps = len(apps[0].per_step)
  if steps <= 0:
    raise ValueError("No per-step results found.")
  for app in apps[1:]:
    if len(app.per_step) != steps:
      raise ValueError(f"step count mismatch: {apps[0].source_dir}={steps} vs {app.source_dir}={len(app.per_step)}")

  args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)

  wb = Workbook()
  wb.remove(wb.active)

  ws_matrix = wb.create_sheet("Matrix")
  write_matrix_sheet(
      ws_matrix,
      apps=apps,
      steps=steps,
      value_for_cell=lambda app, r: "" if r is None else format_step_cell(fixed_policies, r),
      column_width=60,
  )

  ws_best_action = wb.create_sheet("BestAction")
  write_matrix_sheet(
      ws_best_action,
      apps=apps,
      steps=steps,
      value_for_cell=lambda app, r: ""
      if r is None
      else pick_best_action_key(fixed_policies, (r.get("per_action_ipc") or {}))[0],
      column_width=45,
  )

  ws_best_ipc = wb.create_sheet("BestIPC")

  def best_ipc_cell(app: AppSummary, r: Optional[Mapping[str, Any]]):
    if r is None:
      return ""
    best_action, best_ipc = pick_best_action_key(fixed_policies, (r.get("per_action_ipc") or {}))
    return best_ipc

  write_matrix_sheet(
      ws_best_ipc,
      apps=apps,
      steps=steps,
      value_for_cell=best_ipc_cell,
      column_width=12,
  )
  for row in range(2, steps + 2):
    for col in range(2, len(apps) + 2):
      ws_best_ipc.cell(row=row, column=col).number_format = "0.000000"

  ws_long = wb.create_sheet("AllIPC")
  ws_long.append(["application", "step", "action", "ipc", "is_best", "best_action", "best_ipc", "base_ipc"])
  ws_long.freeze_panes = "A2"
  ws_long.column_dimensions["A"].width = 24
  ws_long.column_dimensions["B"].width = 6
  ws_long.column_dimensions["C"].width = 60
  ws_long.column_dimensions["D"].width = 12
  ws_long.column_dimensions["E"].width = 8
  ws_long.column_dimensions["F"].width = 60
  ws_long.column_dimensions["G"].width = 12
  ws_long.column_dimensions["H"].width = 12
  for cell in ws_long[1]:
    cell.font = Font(bold=True)

  # Fill long sheet
  for app in apps:
    by_step: Dict[int, Mapping[str, Any]] = {}
    for r in app.per_step:
      step = r.get("step")
      if isinstance(step, int):
        by_step[step] = r
    for step in range(steps):
      r = by_step.get(step)
      if r is None:
        continue
      per_action_ipc = r.get("per_action_ipc") or {}
      if not isinstance(per_action_ipc, dict):
        per_action_ipc = {}

      best_action, best_ipc = pick_best_action_key(fixed_policies, per_action_ipc)
      base_ipc = r.get("base_ipc")
      for action in fixed_policies:
        ws_long.append(
            [
                app.app,
                step,
                action,
                per_action_ipc.get(action),
                1 if best_action == action else 0,
                best_action,
                best_ipc,
                base_ipc,
            ]
        )

  wb.save(args.out_xlsx)

  if not args.skip_csv:
    write_long_csv(args.out_long_csv, apps, steps=steps, fixed_policies=fixed_policies)

  print(f"Wrote: {args.out_xlsx}")
  if not args.skip_csv:
    print(f"Wrote: {args.out_long_csv}")


if __name__ == "__main__":
  main()
